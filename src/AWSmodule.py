import openpyxl as excel
import os
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException


'''
main                
|-  input_contacts(db_col)         
|-  input_message(campaign)            
|-  sender()            
    |-  send_unsaved_contact_message()      
    |-  send_attachment()       
    |-  send_files()        
'''

def read_message(fileName):
    msgIklan = open(fileName, 'r').read()

    myText = msgIklan[0]
    for j in range(1, len(msgIklan)):
        myText += msgIklan[j]
    # myText = myText.replace('\n', (Keys.SHIFT + Keys.ENTER + Keys.SHIFT));

    return myText

def read_database(fileName, s_name, col_name):
    lst = []
    workbook = excel.load_workbook(fileName)
    sheet = workbook[s_name]
    firstCol = sheet[col_name]
    for cell in range(len(firstCol)):
        xlist = str(firstCol[cell].value)
        if xlist != 'None':
        # contact = "\"" + contact + "\""
            lst.append(xlist)
    return lst

def read_medialist (medialist, docType):
    # medialist = open(medialist, 'r').read().split('\n')

    a = ''

    # for i in range(len(medialist)):
    for i in range(len(medialist[1:])):
        # image_path = os.getcwd() + "\\Media\\" + medialist[i]
        if docType == 1:
            image_path = os.getcwd() + "\\Media\\" + medialist[i+1]
        else:
            image_path = os.getcwd() + "\\Documents\\" + medialist[i+1]
        # image_path = image_path.replace("\\", "\\\\")
        image_path = image_path.replace('\\src', '')
        image_path = '"' + image_path + '" '
        a += image_path 

    return a

def read_setting(fileName, s_name='setting', col_name=2):
    lst = []
    workbook = excel.load_workbook(fileName)
    sheet = workbook[s_name]

    i = 4
    a = sheet.cell(row=i+1,column=col_name).value
    b = sheet.cell(row=i+2,column=col_name).value
    c = sheet.cell(row=i+3,column=col_name).value
    d = sheet.cell(row=i+4,column=col_name).value

    return a, b, c, d

def webloading(browser, link, address_XPATH):    
    browser.get(link)
    timeout = 3 # seconds
    i = 1
    while True:
        try:
            element_present = EC.presence_of_element_located((By.XPATH, address_XPATH))
            WebDriverWait(browser, timeout).until(element_present)
            validation = True
            break
        except TimeoutException:
            try:
                aa = browser.find_element_by_xpath('//*[@id="app"]/div/span[2]/div/span/div/div/div/div/div/div[1]')
                validation = False
                break
            except NoSuchElementException:
                pass
            # if not aa:
                # print("No element found") 
            #     print(aa)
            # else:
            #     # element = elements[0]
            #     print("element found")
            #     print(aa)
            print (i, "Waiting for web page loading...")
            i+=1
    return validation
def autoitloading(autoit):    
    timeout = 3 # seconds
    i = 1
    while True:
        try:
            autoit.win_wait_active("Open", timeout)
            break
        except TimeoutException:
            print (i, "Waiting for window apprear...")
            i+=1


if __name__ == '__main__':
    # wb_name = 'database.xlsx'
    # medialist = read_database(wb_name, 'madu', 'D')


    # image_path = read_medialist(medialist)
    # print(image_path)
    # link_num = "https://web.whatsapp.com/send?phone={}&text&source&data&app_absent".format(6282210138809)
    # validatorsnumber(link_num)
    # validatorsnumber(link_num)
    pass